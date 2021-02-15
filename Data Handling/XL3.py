import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet_title = sheet.title
print(f"Sheet name: {sheet_title}")
sheet.title = "Accounts"
print(f"Sheet is renmaed as: {sheet.title}")
c1 = sheet.cell(row = 1, column = 1)
c1.value = "RST"
c2 = sheet.cell(row = 1, column = 2)
c2.value = "Forum"
c3 = sheet["A3"]
c3.value = "Piyush"
c4 = sheet["B2"]
c4.value = "Abhilash"
wb.save("C:\\Users\\Dell\\OneDrive\\Desktop\\XL3.xlsx")