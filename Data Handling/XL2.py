import openpyxl

path = "C:\\Users\\Dell\\OneDrive\\Desktop\\RST_Forum.xlsx"
wb_obj = openpyxl.load_workbook(path)
sh_obj = wb_obj.active
cell_obj = sh_obj.cell(row = 1, column = 1)
print(f"Value of the first cell is {cell_obj.value}")
print(f"Total Rows are: {sh_obj.max_row}")
print(f"Total columns are: {sh_obj.max_column}")
for i in range(1, sh_obj.max_column + 1):
    cell_obj = sh_obj.cell(row = 1, column = i)
    print(f"Names of the column are: {cell_obj.value}")
for j in range(1, sh_obj.max_row + 1):
    cell_obj = sh_obj.cell(row = j, column = 2)
    print(f"The name are {cell_obj.value}")
for i in range(1, sh_obj.max_column + 1):
    cell_obj = sh_obj.cell(row = 2, column = i)
    print(cell_obj.value, end="-")