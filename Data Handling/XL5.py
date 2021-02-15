import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells("A1:D4")
sheet.cell(row = 1, column = 1).value = "12 cells nerged togather"
sheet.merge_cells("C6:D6")
sheet.cell(row = 3, column = 6).value = "2cells"
wb.save("C:\\Users\\Dell\\OneDrive\\Desktop\\XL5.xlsx")