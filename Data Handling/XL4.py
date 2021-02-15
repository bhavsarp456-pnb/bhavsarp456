import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet.cell(row = 1, column = 1).value = "Sunny"
sheet.cell(row = 1, column = 2).value = "Deol"
sheet.row_dimensions[1].height = 70
sheet.column_dimensions["B"].width = 20
wb.save("C:\\Users\\Dell\\OneDrive\\Desktop\\XL4.xlsx")