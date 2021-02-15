from openpyxl import load_workbook, Workbook

mywb = load_workbook(
    "C:\\Users\\Dell\\OneDrive\\Desktop\\Sales.xlsx"
)
mywb = Workbook()
mywb.sheetnames
mywb.create_sheet("Operations", 2)
mysh = mywb["Operations"]
mycell = mysh["C8"]
mycell.value
mycell.value = "Piyush"
myblock = mysh["D3:F12"]
for eachrow in myblock:
    for eachcell in eachrow:
        eachcell.value = "RST Forum"
mywb.save("Sales_up.xlsx")
