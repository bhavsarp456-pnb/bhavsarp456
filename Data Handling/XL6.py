import openpyxl
path = "C:\\Users\\Dell\\OneDrive\\Desktop\\XL5.xlsx"

wb = openpyxl.Workbook(path)
sheet = wb.active
sheet.unmerge_cells("A1:D4")
sheet.cell(row = 1, column = 1).value = "12 cells unmerged individually"
sheet.unmerge_cells("C6:D6")
sheet.cell(row = 3, column = 6).value = "2cells"
path1 = "C:\\Users\\Dell\\OneDrive\\Desktop\\XL6.xlsx"
wb.save(path1)
