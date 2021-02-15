from openpyxl import load_workbook, Workbook

mywb = load_workbook(
    "C:\\Users\\Dell\\OneDrive\\Desktop\\Piyush_Sheet.xlsx"
)
mywb = Workbook()
mywb.sheetnames
mywb.create_sheet("Operations", 2)
mysh = mywb["Operations"]
mycell = mysh["B7"]
mycell.value
mycell.value = "Paresh"
myblock = mysh["F1:H10"]
for eachrow in myblock:
    for eachcell in eachrow:
        eachcell.value = "Paresh"
mywb.save("Piyush_sheet.xlsx")

